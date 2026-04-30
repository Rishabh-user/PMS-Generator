"""
Generates formatted Excel PMS sheets matching the standard industrial format.
Uses openpyxl for precise formatting control.
"""
import io
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.pms_models import PMSResponse

# Logo banner bounds (cols A:E × rows 1:6 — the full merged logo cell).
#   Cols A=16, B=11, C=11, D=11, E=11 char widths ≈ 60 chars × ~7 px ≈ 420 px usable.
#   Rows 1-3 are 22 pt each, rows 4-6 are 20 pt each → 126 pt ≈ 168 px usable.
#   (Row 5 may grow with long service strings; the few-pixel drift is fine
#    because the logo is much shorter than the banner and stays centred.)
# Image is scaled UNIFORMLY by whichever dimension is binding so it fills
# the banner without overflowing into column F / row 7. Original aspect
# ratio is preserved, and the image is then centred horizontally and
# vertically inside the full 6-row banner via offset on a OneCellAnchor.
LOGO_MAX_WIDTH_PX  = 380
LOGO_MAX_HEIGHT_PX = 150
BANNER_WIDTH_PX    = int((16 + 11 + 11 + 11 + 11) * 7)         # ≈ 420 px
BANNER_HEIGHT_PX   = int((22 + 22 + 22 + 20 + 20 + 20) * 4 / 3)  # = 168 px
PX_TO_EMU          = 9525                                       # 1 px @ 96 DPI

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
DATA_FILL = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
NOTES_FILL = PatternFill("solid", fgColor="FFF2CC")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="1F4E79", size=14)
SECTION_FONT = Font(name="Arial", bold=True, color="1F4E79", size=10)
LABEL_FONT = Font(name="Arial", bold=True, color="333333", size=9)
DATA_FONT = Font(name="Arial", color="333333", size=9)
NOTE_FONT = Font(name="Arial", italic=True, color="666666", size=8)

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="1F4E79"))

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _logo_bytes() -> bytes | None:
    """Return the logo PNG bytes from app/static/images/.

    Tries `excel-logo.png` first (override slot for a purpose-built variant),
    then falls back to the default `logo.png`. Returns None if neither file
    exists — spec generation continues without branding rather than failing.
    """
    base = Path(__file__).resolve().parent.parent / "static" / "images"
    for name in ("excel-logo.png", "logo.png"):
        p = base / name
        try:
            if p.exists():
                return p.read_bytes()
        except OSError as e:
            logger.warning("Logo path %s unreadable: %s", p, e)
    return None


def _insert_logo(ws, anchor_cell: str = "A1") -> None:
    """Embed the project logo in the top-left banner area (cells A1:E6).

    Image is proportionally resized to fit LOGO_MAX_WIDTH_PX × LOGO_MAX_HEIGHT_PX
    so the original aspect ratio is preserved, then centred horizontally
    AND vertically inside the full 6-row × 5-col merged banner cell.
    Silently no-ops if the logo bytes can't be obtained — spec generation
    must keep working even without branding.
    """
    data = _logo_bytes()
    if not data:
        logger.warning("Logo bytes unavailable — skipping insertion")
        return
    try:
        img = XLImage(io.BytesIO(data))
        # Scale uniformly to fit the banner bounding box — whichever dimension
        # is more constraining wins, so the image never overflows into the
        # header table on the right or the rows below.
        if img.height and img.width:
            scale = min(LOGO_MAX_WIDTH_PX / img.width, LOGO_MAX_HEIGHT_PX / img.height)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)

        # Centre horizontally + vertically inside the A1:C3 banner. We attach
        # a custom OneCellAnchor whose colOff / rowOff push the image away
        # from A1's top-left corner by half the leftover banner space.
        x_off_px = max(0, (BANNER_WIDTH_PX  - img.width)  // 2)
        y_off_px = max(0, (BANNER_HEIGHT_PX - img.height) // 2)
        marker = AnchorMarker(
            col=0, colOff=x_off_px * PX_TO_EMU,
            row=0, rowOff=y_off_px * PX_TO_EMU,
        )
        ext = XDRPositiveSize2D(
            cx=img.width  * PX_TO_EMU,
            cy=img.height * PX_TO_EMU,
        )
        img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws.add_image(img)
        logger.info(
            "Logo centred at A1 (offset %d×%d px, size %d×%d px, banner %d×%d)",
            x_off_px, y_off_px, img.width, img.height, BANNER_WIDTH_PX, BANNER_HEIGHT_PX,
        )
    except Exception as e:
        logger.warning("Failed to insert logo: %s", e)


def _short_schedule(sched) -> str:
    """Render schedule for the compact pipe-data row.

    Strips the leading 'SCH ' prefix so the cell shows '160' instead of
    'SCH 160' (matches the project's reference layout). Bare schedule
    names like 'STD' / 'XS' / 'XXS' / '5S' / '40S' / '-' pass through
    unchanged.
    """
    if sched is None:
        return ""
    s = str(sched).strip()
    upper = s.upper()
    for prefix in ("SCHEDULE ", "SCH ", "SCH"):
        if upper.startswith(prefix):
            return s[len(prefix):].strip() or s
    return s


def _get_sheet_no(piping_class: str) -> str:
    """Return the sheet number shown in the header table.

    Preference order:
      1. `sheet_no` field on the catalogue entry (if curated there later).
      2. 1-based position of the class in `pipe_classes.json`.
      3. Empty string as a safe fallback.
    """
    try:
        from app.services import data_service
        entries = data_service.get_all_entries()
        target = (piping_class or "").strip().upper()
        for idx, e in enumerate(entries, start=1):
            if e.get("piping_class", "").strip().upper() == target:
                explicit = e.get("sheet_no")
                if explicit:
                    return str(explicit)
                return str(idx)
    except Exception:
        pass
    return ""


def _write_pms_header(ws, pms: PMSResponse, total_cols: int) -> int:
    """Write the 6-row header banner and return the next free row.

    Layout — single big merged logo on the left (A1:E6) plus the full
    header content stacked on the right (cols F..total_cols), matching
    the reference PMS document:

        ┌─────────────┬───────────────────────────────────┬────────┐
    Row1│             │     PIPING MATERIAL SPECIFICATION │ Rev: A0│
        │             ├─────────┬───────┬─────┬─────┬─────┤        │
    Row2│             │ Piping  │Material│ C.A │Mill │Sheet│        │
        │             │  Class  │       │     │ Tol │ No. │        │
    Row3│    LOGO     ├──┬──────┼───────┼─────┼─────┼─────┤        │
        │             │A1│ 150# │  CS   │3 mm │12.5%│  27 │        │
        │             ├──┴──────┴───────┴─────┴─────┴─────┴────────┤
    Row4│             │ Design Code: │       ASME B 31.3            │
    Row5│             │ Service:     │ Cooling Media, Heating, …    │
    Row6│             │ Branch Chart:│ Ref. APPENDIX-1, Chart 1     │
        └─────────────┴──────────────────────────────────────────────┘

    Column groups on the right (proportional to available width):
      Piping Class : 3 cols out of 18  (Class value spans 2, Rating 1)
      Material     : 6 cols
      C.A          : 3 cols
      Mill Tol     : 2 cols
      Sheet No.    : 4 cols
      Title row uses the same 18 cols; Rev sits at the rightmost 3.
    """
    # ── Column widths ──
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    # The right-side header cols default to ~10; pipe-data cols K+ stay at
    # width 6 (set by the caller). No special per-letter widths here — the
    # cell merges define the visual blocks.

    LOGO_COL_END = 5            # logo banner spans cols 1-5 (A:E)
    RIGHT_COL_START = 6         # header table starts at col F

    # ── Compute the proportional column groups for the right side ──
    # Available width on the right: total_cols - 5 = 18 cols for A1.
    # Distribute by weight: PC 3 / MAT 6 / CA 3 / MT 2 / SN 4 = 18.
    right_w = max(total_cols - LOGO_COL_END, 14)  # at least 14 cols
    weights = {"pc": 3, "mat": 6, "ca": 3, "mt": 2, "sn": 4}
    total_weight = sum(weights.values())  # 18
    widths = {k: max(2, round(w * right_w / total_weight)) for k, w in weights.items()}
    # Last group absorbs any rounding remainder
    widths["sn"] = right_w - sum(v for k, v in widths.items() if k != "sn")

    # Column boundaries (start..end inclusive) for each header group.
    pc_s = RIGHT_COL_START
    pc_e = pc_s + widths["pc"] - 1
    mat_s = pc_e + 1
    mat_e = mat_s + widths["mat"] - 1
    ca_s = mat_e + 1
    ca_e = ca_s + widths["ca"] - 1
    mt_s = ca_e + 1
    mt_e = mt_s + widths["mt"] - 1
    sn_s = mt_e + 1
    sn_e = sn_s + widths["sn"] - 1
    # Sub-split inside Piping Class group: class value spans first 2 cols,
    # rating value gets the 3rd. (Both share Header label.)
    pc_class_e = pc_s + max(1, widths["pc"] - 2)  # class spans up to here
    pc_rating_s = pc_class_e + 1                  # rating starts here

    # Title row 1: title fills the left portion; Rev fits on the right.
    # Reuse the Sheet No. group for Rev so the column ratios align.
    title_s = RIGHT_COL_START
    title_e = mt_e        # ends just before Sheet No. group
    rev_s = sn_s          # Rev sits in Sheet No.'s columns on row 1
    rev_e = sn_e

    # ── Row heights ──
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20
    # Row 5 height is set later based on the service string length.
    ws.row_dimensions[6].height = 20

    # ── Logo block: one big merged cell A1:E6 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=6, end_column=LOGO_COL_END)
    logo_cell = _apply_style(
        ws, 1, 1, font=DATA_FONT, fill=DATA_FILL,
        alignment=CENTER, border=THIN_BORDER,
    )
    logo_cell.value = ""    # placeholder — image overlays this
    _insert_logo(ws, "A1")

    # ── Row 1: Title + Rev ──
    ws.merge_cells(start_row=1, start_column=title_s, end_row=1, end_column=title_e)
    title_cell = _apply_style(
        ws, 1, title_s,
        font=Font(name="Arial", bold=True, color="1F4E79", size=13),
        fill=ALT_FILL, alignment=CENTER, border=THIN_BORDER,
    )
    title_cell.value = "PIPING MATERIAL SPECIFICATION"
    for c in range(title_s, title_e + 1):
        _apply_style(ws, 1, c, font=DATA_FONT, fill=ALT_FILL, border=THIN_BORDER)
    # Rev block (right edge, narrower)
    ws.merge_cells(start_row=1, start_column=rev_s, end_row=1, end_column=rev_e)
    rev_cell = _apply_style(
        ws, 1, rev_s,
        font=Font(name="Arial", bold=True, color="333333", size=10),
        fill=ALT_FILL, alignment=CENTER, border=THIN_BORDER,
    )
    rev_cell.value = f"Rev: {pms.version or 'A0'}"
    for c in range(rev_s, rev_e + 1):
        _apply_style(ws, 1, c, font=DATA_FONT, fill=ALT_FILL, border=THIN_BORDER)

    # ── Row 2: Column headers (Piping Class / Material / C.A / Mill Tol / Sheet No.) ──
    header_groups = [
        (pc_s,  pc_e,  "Piping Class"),
        (mat_s, mat_e, "Material"),
        (ca_s,  ca_e,  "C.A"),
        (mt_s,  mt_e,  "Mill Tol"),
        (sn_s,  sn_e,  "Sheet No."),
    ]
    for s, e, label in header_groups:
        if s != e:
            ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        for c in range(s, e + 1):
            _apply_style(ws, 2, c, font=LABEL_FONT, fill=ALT_FILL,
                         alignment=CENTER, border=THIN_BORDER)
        ws.cell(row=2, column=s).value = label

    # ── Row 3: Values ──
    # Piping Class group is split: class value spans 2 cols, rating spans 1
    if pc_class_e >= pc_s and pc_class_e > pc_s:
        ws.merge_cells(start_row=3, start_column=pc_s, end_row=3, end_column=pc_class_e)
    if pc_rating_s <= pc_e and pc_rating_s != pc_e:
        ws.merge_cells(start_row=3, start_column=pc_rating_s, end_row=3, end_column=pc_e)
    for c in range(pc_s, pc_e + 1):
        _apply_style(ws, 3, c, font=DATA_FONT, fill=DATA_FILL,
                     alignment=CENTER, border=THIN_BORDER)
    ws.cell(row=3, column=pc_s).value = pms.piping_class
    ws.cell(row=3, column=pc_rating_s).value = pms.rating

    value_groups = [
        (mat_s, mat_e, pms.material),
        (ca_s,  ca_e,  pms.corrosion_allowance),
        (mt_s,  mt_e,  pms.mill_tolerance),
        (sn_s,  sn_e,  _get_sheet_no(pms.piping_class)),
    ]
    for s, e, value in value_groups:
        if s != e:
            ws.merge_cells(start_row=3, start_column=s, end_row=3, end_column=e)
        for c in range(s, e + 1):
            _apply_style(ws, 3, c, font=DATA_FONT, fill=DATA_FILL,
                         alignment=CENTER, border=THIN_BORDER)
        ws.cell(row=3, column=s).value = value

    # ── Rows 4-6: Design Code / Service / Branch Chart ──
    # Label spans the Piping Class group width (cols F..H ≈ 3 cols),
    # value spans everything else on the right (Material..Sheet No.).
    label_s = RIGHT_COL_START
    label_e = pc_e
    val_s = mat_s
    val_e = sn_e
    full_rows = [
        ("Design Code:", pms.design_code or ""),
        ("Service:", pms.service or ""),
        ("Branch Chart:", pms.branch_chart or ""),
    ]
    for i, (label, value) in enumerate(full_rows):
        r = 4 + i

        # Label side
        if label_s != label_e:
            ws.merge_cells(start_row=r, start_column=label_s, end_row=r, end_column=label_e)
        for c in range(label_s, label_e + 1):
            _apply_style(ws, r, c, font=LABEL_FONT, fill=ALT_FILL,
                         alignment=LEFT, border=THIN_BORDER)
        ws.cell(row=r, column=label_s).value = label

        # Value side (left-aligned, spans Material..Sheet No.)
        if val_s != val_e:
            ws.merge_cells(start_row=r, start_column=val_s, end_row=r, end_column=val_e)
        for c in range(val_s, val_e + 1):
            _apply_style(ws, r, c, font=DATA_FONT, fill=DATA_FILL,
                         alignment=LEFT, border=THIN_BORDER)
        ws.cell(row=r, column=val_s).value = value

        # Wrap long service strings — bump row height proportional to length
        if label == "Service:":
            ws.row_dimensions[r].height = max(
                (len(str(value)) // 60 + 1) * 14, 20
            )

    return 7  # next free row


def _apply_style(ws, row, col, font=DATA_FONT, fill=DATA_FILL, alignment=CENTER, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell


def _write_section_header(ws, row: int, text: str, col_start: int = 1, col_end: int = 20):
    """Write a section header row (merged and centered across the full width)."""
    for c in range(col_start, col_end + 1):
        _apply_style(ws, row, c, font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER)
    ws.cell(row=row, column=col_start).value = text
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)


def _valve_row_has_data(fallback: str, by_size, pipe_sizes: list) -> bool:
    """Decide whether a valve row should render in the Excel sheet.

    Returns True if the row has SOMETHING to show, namely either:
      • the class-level fallback string is non-blank, OR
      • the per-size list has at least one entry whose code is non-blank
        AND whose size_inch matches one of the pipe sizes the class supports
        (so a stale entry for an unused size doesn't keep an otherwise
        empty row visible).

    Used to suppress Butterfly / DBB / DBB (Inst) rows on classes / sizes
    where those valves aren't applicable, instead of showing an empty band.
    """
    if (fallback or "").strip():
        return True
    sizes_str = {str(s) for s in (pipe_sizes or [])}
    for entry in (by_size or []):
        code = (getattr(entry, "code", "") or "").strip()
        sz = str(getattr(entry, "size_inch", "") or "")
        if code and (not sizes_str or sz in sizes_str):
            return True
    return False


def _lvcf_by_size(by_size, pipe_sizes: list) -> list:
    """Last-Value-Carried-Forward: expand a sparse `{size: code}` list from
    the AI (e.g. only boundary sizes given) to a full list aligned with
    `pipe_sizes`, so the merge logic renders contiguous groups correctly.

    Example (A1 Check):
        by_size    = [{"0.5": "CHPMA1R"}, {"2": "CHSMA1R, CHDMA1R"}]
        pipe_sizes = ["0.5","0.75","1","1.5","2","3","4","6"...]
        result     = ["CHPMA1R","CHPMA1R","CHPMA1R","CHPMA1R",
                      "CHSMA1R, CHDMA1R","CHSMA1R, CHDMA1R", …]

    Entries are sorted by numeric size before propagation. Any size in
    pipe_sizes that falls before the first explicit entry gets "".

    Post-filter: "USE GATE VALVE" is a small-bore-only placeholder used on
    the Ball row for E/F/G-series classes where a ball valve isn't
    available at small sizes. Regardless of what boundary the AI emits
    (e.g. LVCF might carry it forward to size 6"), cap this text at
    sizes ≤ 1.5" — any cell > 1.5" carrying "USE GATE VALVE" is blanked.
    """
    def _num(s):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None

    entries = sorted(
        [(_num(e.size_inch), e.code) for e in by_size if _num(e.size_inch) is not None],
        key=lambda x: x[0],
    )
    out = []
    current = ""
    idx = 0
    for size in pipe_sizes:
        tgt = _num(size)
        if tgt is None:
            out.append(current)
            continue
        while idx < len(entries) and entries[idx][0] <= tgt:
            current = entries[idx][1]
            idx += 1
        out.append(current)

    # Cap "USE GATE VALVE" at 1.5" — project rule.
    for i, (size, val) in enumerate(zip(pipe_sizes, out)):
        if not val:
            continue
        if "USE GATE VALVE" not in val.upper():
            continue
        n = _num(size)
        if n is not None and n > 1.5:
            out[i] = ""
    return out


def _write_merged_data_row(ws, row: int, label: str, values: list, col_start: int = 2,
                           total_cols: int = 20, font=DATA_FONT, fill=DATA_FILL):
    """Write a data row that auto-merges consecutive cells with identical values.

    Layout: col 1 = label, col_start..total_cols = data columns.
    """
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=fill, alignment=LEFT).value = label
    for i in range(len(values)):
        col = col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=font, fill=fill, alignment=CENTER)
    for c in range(col_start + len(values), total_cols + 1):
        _apply_style(ws, row, c, fill=fill)

    if not values:
        return

    run_start = 0
    for i in range(1, len(values) + 1):
        if i == len(values) or str(values[i]) != str(values[run_start]):
            start_col = col_start + run_start
            end_col = col_start + i - 1
            if start_col <= total_cols:
                end_col = min(end_col, total_cols)
                ws.cell(row=row, column=start_col).value = values[run_start]
                if end_col > start_col:
                    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            if i < len(values):
                run_start = i


def _write_size_header_row(ws, row: int, sizes: list, col_start: int = 2, total_cols: int = 20):
    """Write a size header row: col 1 = 'Size (in)', col_start..total_cols = size values.

    The row is HIDDEN by default — only Pipe Data shows the size labels
    visually. All other sections (Fittings, Flange, Spectacle Blind, Bolts /
    Nuts / Gaskets, Valves) keep the column-to-size mapping for layout
    correctness but suppress the redundant header so the spec reads cleaner.
    """
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Size (in)"
    for i, size in enumerate(sizes):
        col = col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=LABEL_FONT, fill=ALT_FILL, alignment=CENTER).value = size
    for c in range(col_start + len(sizes), total_cols + 1):
        _apply_style(ws, row, c, fill=ALT_FILL)
    ws.row_dimensions[row].hidden = True


def _write_label_offset_value_row(ws, row: int, label: str, value: str, value_start_col: int, col_end: int, fill=DATA_FILL):
    """Col 1 = label; cols 2..value_start_col-1 = blank fill; value_start_col..col_end = merged value.

    Used for Compact Flange / Hub Connector rows where the value only applies to larger sizes.
    """
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=fill, alignment=LEFT).value = label
    for c in range(2, value_start_col):
        _apply_style(ws, row, c, fill=fill)
    _apply_style(ws, row, value_start_col, font=DATA_FONT, fill=fill, alignment=LEFT).value = value
    if col_end > value_start_col:
        ws.merge_cells(start_row=row, start_column=value_start_col, end_row=row, end_column=col_end)
    for c in range(value_start_col + 1, col_end + 1):
        _apply_style(ws, row, c, fill=fill)


def _write_range_value_row(ws, row: int, label: str, value: str,
                           start_col: int, end_col: int, total_cols: int, fill=DATA_FILL):
    """Label in col 1; all data columns blank-filled; value merged across [start_col..end_col].

    Used when a row's value applies to a specific pipe-size range only
    (e.g. Plug fittings from 0.5" to 1.5").
    """
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=fill, alignment=LEFT).value = label
    for c in range(2, total_cols + 1):
        _apply_style(ws, row, c, fill=fill)
    if not value:
        return
    start_col = max(start_col, 2)
    end_col = min(end_col, total_cols)
    if end_col < start_col:
        return
    _apply_style(ws, row, start_col, font=DATA_FONT, fill=fill, alignment=CENTER).value = value
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def _size_column_index(pipe_sizes: list, target_size: str, pipe_col_start: int = 2, default_offset: int = 5) -> int:
    """Return the Excel column index for the given pipe size (e.g. '3'). Falls back to default_offset if not found."""
    for i, s in enumerate(pipe_sizes):
        if str(s).strip().rstrip('"') == target_size:
            return pipe_col_start + i
    return pipe_col_start + default_offset


# ASME B 16.48 project size boundary for Spectacle Blinds.
# Sizes ≤ this value use "ASME B 16.48"; sizes > this use "Spacer and blind...".
# Previously split was index-based (len // 2), which silently shifted the
# boundary when the class's size count changed. Using a size cutoff keeps
# the boundary stable. Per verified Excel spec:
#   SS 316L 10-series (A10, A10N, etc.): boundary = 12" (Spacer from 14")
#   All other classes:                   boundary = 14" (Spacer from 16")
B1648_MAX_SIZE_DEFAULT = 14.0
B1648_MAX_SIZE_SS_10SERIES = 12.0


def _b1648_max_size_for(pms) -> float:
    """Return the project-specific B16.48 size cutoff for the class.

    SS 316L 10-series (A10 / A10N, and by extension B10/D10/E10/F10/G10 + N
    variants): cutoff = 12" — Spacer and blind starts at 14".
    Every other class: cutoff = 14" — Spacer and blind starts at 16".
    """
    cls = (getattr(pms, "piping_class", "") or "").upper().strip()
    # Match "A10", "A10N", "B10", "B10N", "D10", ..., "G10N" (10-series SS 316L)
    # Shape: [A|B|D|E|F|G] + "10" + ("" | "N")
    if len(cls) >= 3 and cls[0] in "ABDEFG" and cls[1:3] == "10":
        tail = cls[3:]
        if tail in ("", "N"):
            return B1648_MAX_SIZE_SS_10SERIES
    return B1648_MAX_SIZE_DEFAULT


def _any_nonempty(values) -> bool:
    """True if any value in the list is a non-empty string after stripping."""
    return any(str(v or "").strip() for v in values)


def _split_index_at_size(pipe_data, max_size_inches: float) -> int:
    """Return the index of the first pipe_data entry whose size exceeds
    max_size_inches. If all sizes are ≤ max, returns len(pipe_data).
    If all sizes are > max, returns 0."""
    for i, p in enumerate(pipe_data):
        try:
            size_val = float(str(p.size_inch).strip().rstrip('"'))
        except (ValueError, TypeError, AttributeError):
            continue
        if size_val > max_size_inches:
            return i
    return len(pipe_data)


def _render_spectacle_blind_row(
    ws, row: int, pms, pipe_col_start: int, total_cols: int,
) -> None:
    """Render the Spectacle row with a size-based split at 14" (B1648_MAX_SIZE_INCHES).
    Handles three cases: all small / all large / mixed."""
    for c in range(1, total_cols + 1):
        _apply_style(ws, row, c, font=DATA_FONT, fill=ALT_FILL, alignment=CENTER)
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Spectacle"

    standard = pms.spectacle_blind.standard or ""
    standard_large = pms.spectacle_blind.standard_large or ""
    n = len(pms.pipe_data)
    if n == 0:
        return

    cutoff = _b1648_max_size_for(pms)
    split = _split_index_at_size(pms.pipe_data, cutoff) if standard_large else n

    if split == 0 and standard_large:
        # All sizes > 14": only standard_large applies
        ws.merge_cells(start_row=row, start_column=pipe_col_start, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=pipe_col_start).value = standard_large
    elif split >= n or not standard_large:
        # All sizes ≤ 14" (or no large-standard defined)
        ws.merge_cells(start_row=row, start_column=pipe_col_start, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=pipe_col_start).value = standard
    else:
        left_end = pipe_col_start + split - 1
        right_start = pipe_col_start + split
        ws.merge_cells(start_row=row, start_column=pipe_col_start, end_row=row, end_column=left_end)
        ws.cell(row=row, column=pipe_col_start).value = standard
        ws.merge_cells(start_row=row, start_column=right_start, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=right_start).value = standard_large


def _write_label_value_row(ws, row: int, label: str, value: str, col_start: int = 1, val_col: int = 2, col_end: int = 20):
    """Write a label-value pair row.

    Layout:
      Col 1 (A) = label
      Cols val_col..col_end (B onward, merged) = value
    """
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=DATA_FILL, alignment=LEFT).value = label
    _apply_style(ws, row, val_col, font=DATA_FONT, fill=DATA_FILL, alignment=LEFT).value = value
    if col_end > val_col:
        ws.merge_cells(start_row=row, start_column=val_col, end_row=row, end_column=col_end)
    for c in range(val_col + 1, col_end + 1):
        _apply_style(ws, row, c, fill=DATA_FILL)


def generate_pms_excel(pms: PMSResponse, output_path: Path) -> Path:
    """Generate a formatted Excel PMS sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = pms.piping_class

    # Determine column count based on pipe sizes
    num_pipe_cols = max(len(pms.pipe_data), 1)
    # Col 1 = label, col 2..(N+1) = size-data. No trailing empty column —
    # the title-row banner extends to the rightmost data col via fill.
    total_cols = max(num_pipe_cols + 1, 19)
    pipe_col_start = 2

    # Per-pipe-size column widths — compact format per Image-2 reference.
    # Cols A-J are sized by the header helper; cols K+ carry the size-by-size
    # data and are kept narrow (~6) so the full 22-size row fits without
    # horizontal scrolling.
    for i in range(11, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 6

    # === HEADER BANNER (rows 1-6): compact spec table ===
    row = _write_pms_header(ws, pms, total_cols)

    # === PRESSURE-TEMPERATURE RATING ===
    _write_section_header(ws, row, "Pressure-Temperature Rating", col_end=total_cols)
    row += 1

    pt = pms.pressure_temperature

    # Pressure row (compact label, comma form per the reference sheet)
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=DATA_FILL, alignment=LEFT).value = "Press., barg"
    for i, press in enumerate(pt.pressures):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=DATA_FONT, fill=DATA_FILL).value = press
    # Hydrotest header — single-line, sits to the right of the P/T columns
    ht_col = pipe_col_start + len(pt.temperatures) + 1
    if pms.hydrotest_pressure and ht_col <= total_cols:
        _apply_style(ws, row, ht_col, font=LABEL_FONT, fill=ALT_FILL, alignment=CENTER).value = "Hydrotest Pr. (barg)"
    row += 1

    # Temperature row — single combined row using the labels (e.g. "-29 to 38")
    # so the header doesn't carry a duplicate "Temp. Range" row.
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Temp., °C"
    temp_values = pt.temp_labels if pt.temp_labels else [str(t) for t in pt.temperatures]
    for i, lbl in enumerate(temp_values):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=DATA_FONT, fill=ALT_FILL).value = lbl
    # Hydrotest value beneath its header
    if pms.hydrotest_pressure and ht_col <= total_cols:
        _apply_style(ws, row, ht_col, font=DATA_FONT, fill=DATA_FILL, alignment=CENTER).value = pms.hydrotest_pressure
    row += 1

    # === PIPE DATA ===
    _write_section_header(ws, row, "Pipe Data", col_end=total_cols)
    row += 1

    if pms.pipe_code:
        _write_label_value_row(ws, row, "Code", pms.pipe_code, col_end=total_cols)
        row += 1

    # I.D. (mm) is rendered only when at least one size carries a non-zero
    # inside diameter — used by GRE classes (A50/A51/A52). Row is inserted
    # right after O.D.; other classes skip it.
    pipe_rows = [
        ("Size(in)", [p.size_inch for p in pms.pipe_data]),
        ("O.D.mm", [p.od_mm for p in pms.pipe_data]),
    ]
    if any((getattr(p, "id_mm", 0) or 0) > 0 for p in pms.pipe_data):
        pipe_rows.append(("I.D.mm", [p.id_mm for p in pms.pipe_data]))
    pipe_rows.extend([
        ("Sch.", [_short_schedule(p.schedule) for p in pms.pipe_data]),
        ("WT.mm", [p.wall_thickness_mm for p in pms.pipe_data]),
        ("Type", [p.pipe_type for p in pms.pipe_data]),
        ("MOC", [p.material_spec for p in pms.pipe_data]),
        ("Ends", [p.ends for p in pms.pipe_data]),
    ])

    # Rows that benefit from merging (repeated values across sizes)
    mergeable_pipe_rows = {"Type", "MOC", "Ends"}
    for label, values in pipe_rows:
        is_alt = label in ("Size(in)", "Sch.", "Type", "Ends")
        fill = ALT_FILL if is_alt else DATA_FILL
        if label in mergeable_pipe_rows:
            _write_merged_data_row(ws, row, label, values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        else:
            _apply_style(ws, row, 1, font=LABEL_FONT, fill=fill, alignment=LEFT).value = label
            for i, val in enumerate(values):
                col = pipe_col_start + i
                if col <= total_cols:
                    _apply_style(ws, row, col, font=DATA_FONT, fill=fill).value = val
            for c in range(pipe_col_start + len(values), total_cols + 1):
                _apply_style(ws, row, c, fill=fill)
        row += 1

    # === FITTINGS (SIZE-WISE DATA) ===
    _write_section_header(ws, row, "Fittings — Butt Weld (SCH to match pipe)", col_end=total_cols)
    row += 1

    # Size columns header — written for column-mapping correctness, then
    # hidden because per spec only Pipe Data shows the size labels visually.
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Size (in)"
    for i, fitting in enumerate(pms.fittings_by_size):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=LABEL_FONT, fill=ALT_FILL, alignment=CENTER).value = fitting.size_inch
    for c in range(pipe_col_start + len(pms.fittings_by_size), total_cols + 1):
        _apply_style(ws, row, c, fill=ALT_FILL)
    ws.row_dimensions[row].hidden = True
    row += 1

    # Type row — full descriptive text from AI (e.g. "Butt Weld (SCH to match pipe), Seamless")
    type_values = [f.fitting_type or "" for f in pms.fittings_by_size]
    _write_merged_data_row(ws, row, "Type", type_values, col_start=pipe_col_start,
                           total_cols=total_cols, fill=DATA_FILL)
    row += 1

    # Rating row — GRE classes (A50/A52) show "20 bar, 93degC"; other classes
    # leave fittings.rating empty → row is skipped.
    if (getattr(pms.fittings, "rating", "") or "").strip():
        _write_label_value_row(ws, row, "Rating", pms.fittings.rating, col_end=total_cols)
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    # Fitting properties (MOC, Elbow, Tee, etc.) — with auto-merge.
    # Plug behavior: if the class populates plug_standard for sizes > 1.5"
    # (e.g. Copper A40 with Brazed/BW split), render Plug as a full merged
    # row. If Plug is populated only for small-bore (CS classes where the
    # hex-head threaded plug is small-bore only), render as a range row
    # spanning 0.5"-1.5" only (old behavior preserved).
    # Extra rows (Coupl, Union, Sockolet, Nipple, Swage, plus GRE-specific
    # Mold. Tee / Red. Sad / Adaptor) render only if at least one size has
    # a non-empty value — classes that don't populate these skip the rows.
    fitting_props = [
        ("MOC", lambda f: f.material_spec),
        ("Elbow", lambda f: f.elbow_standard),
        ("Tee", lambda f: f.tee_standard),
        ("Mold. Tee", lambda f: f.mold_tee_standard),
        ("Red.", lambda f: f.reducer_standard),
        ("Red. Sad", lambda f: f.red_saddle_standard),
        ("Cap", lambda f: f.cap_standard),
        ("Coupl", lambda f: f.coupling_standard),
        ("Plug", lambda f: f.plug_standard),
        ("Union", lambda f: f.union_standard),
        ("Sockolet", lambda f: f.sockolet_standard),
        ("Weldolet", lambda f: f.weldolet_spec),
        ("Adaptor", lambda f: f.adaptor_standard),
        ("Nipple", lambda f: f.nipple_standard),
        ("Swage", lambda f: f.swage_standard),
    ]
    # Rows that render only when at least one size carries a value:
    _OPTIONAL_ROWS = {"Coupl", "Union", "Sockolet", "Nipple", "Swage",
                      "Mold. Tee", "Red. Sad", "Adaptor"}

    # Plug is always a small-bore row — see first renderer for rationale.
    fitting_sizes = [f.size_inch for f in pms.fittings_by_size]
    plug_start_col = pipe_col_start
    plug_end_col = _size_column_index(fitting_sizes, "1.5", pipe_col_start=pipe_col_start)

    for prop_idx, (label, getter) in enumerate(fitting_props):
        fill = ALT_FILL if prop_idx % 2 == 0 else DATA_FILL
        prop_values = [getter(f) or "" for f in pms.fittings_by_size]
        if label in _OPTIONAL_ROWS and not _any_nonempty(prop_values):
            continue  # skip row entirely — nothing to show
        if label == "Plug":
            plug_value = ""
            for f, v in zip(pms.fittings_by_size, prop_values):
                if not v:
                    continue
                try:
                    if float(str(f.size_inch).strip().rstrip('"')) <= 1.5:
                        plug_value = v
                        break
                except (ValueError, TypeError):
                    continue
            if not plug_value:
                plug_value = next((v for v in prop_values if v), "")
            _write_range_value_row(ws, row, label, plug_value,
                                   start_col=plug_start_col, end_col=plug_end_col,
                                   total_cols=total_cols, fill=fill)
        else:
            _write_merged_data_row(ws, row, label, prop_values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        row += 1

    pipe_sizes = [p.size_inch for p in pms.pipe_data]

    # === FLANGE DATA ===
    _write_section_header(ws, row, "Flange", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1

    # If compact_flange / hub_connector populated, show WN Flange as label for the Type row
    # and add Compact Flange and Hub Connector rows (reference F1/G1 layout).
    has_extra_flange = bool(pms.flange.compact_flange) or bool(pms.flange.hub_connector)
    type_label = "WN Flange" if has_extra_flange else "Type"
    flange_rows = [
        ("MOC", pms.flange.material_spec, False),
        ("Face", pms.flange.face_type, False),
        (type_label, pms.flange.flange_type, False),
        ("Standard", pms.flange.standard, False),
    ]
    if has_extra_flange:
        flange_rows.append(("Compact Flange", pms.flange.compact_flange, True))
        flange_rows.append(("Hub Connector", pms.flange.hub_connector, True))

    # Compact Flange / Hub Connector values only apply from the 3" size column onward.
    offset_col = _size_column_index(pipe_sizes, "3", pipe_col_start=pipe_col_start)
    for i, (label, value, offset) in enumerate(flange_rows):
        fill = ALT_FILL if i % 2 == 0 else DATA_FILL
        if offset and value:
            _write_label_offset_value_row(ws, row, label, value,
                                          value_start_col=offset_col,
                                          col_end=total_cols, fill=fill)
        else:
            _write_label_value_row(ws, row, label, value, col_end=total_cols)
            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).fill = fill
        row += 1

    # === SPECTACLE BLIND ===
    _write_section_header(ws, row, "Spectacle Blind / Spacer Blinds", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1
    _write_label_value_row(ws, row, "MOC", pms.spectacle_blind.material_spec, col_end=total_cols)
    row += 1
    # Spectacle row — split at 14" (ASME B 16.48 project cutoff)
    _render_spectacle_blind_row(ws, row, pms, pipe_col_start, total_cols)
    ws.row_dimensions[row].height = 25
    row += 1

    # === BOLTS / NUTS / GASKETS ===
    _write_section_header(ws, row, "Bolts / Nuts / Gaskets", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1

    # Washers and Gasket #2 are optional (GRE A50/A51/A52 only); empty →
    # row hidden.
    bng_rows = [
        ("Stud Bolts", pms.bolts_nuts_gaskets.stud_bolts, False),
        ("Hex Nuts", pms.bolts_nuts_gaskets.hex_nuts, False),
        ("Washers", getattr(pms.bolts_nuts_gaskets, "washers", ""), True),
        ("Gasket", pms.bolts_nuts_gaskets.gasket, False),
        ("Gasket", getattr(pms.bolts_nuts_gaskets, "gasket_2", ""), True),
    ]
    rendered = 0
    for label, value, optional in bng_rows:
        if optional and not (value or "").strip():
            continue
        fill = ALT_FILL if rendered % 2 == 0 else DATA_FILL
        _write_label_value_row(ws, row, label, value, col_end=total_cols)
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = fill
        row += 1
        rendered += 1

    # === VALVES ===
    _write_section_header(ws, row, "Valves", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1

    valve_types = [
        ("Rating", pms.valves.rating, []),
        ("Ball", pms.valves.ball, pms.valves.ball_by_size),
        ("Gate", pms.valves.gate, pms.valves.gate_by_size),
        ("Globe", pms.valves.globe, pms.valves.globe_by_size),
        ("Check", pms.valves.check, pms.valves.check_by_size),
        ("Butterfly", pms.valves.butterfly, pms.valves.butterfly_by_size),
        ("DBB (Inst)", pms.valves.dbb_inst, pms.valves.dbb_inst_by_size),
        ("DBB", pms.valves.dbb, pms.valves.dbb_by_size),
    ]
    for i, (label, fallback, by_size) in enumerate(valve_types):
        # Skip rows with no real data so empty Butterfly / DBB / DBB (Inst)
        # bands aren't rendered for classes / sizes where they don't apply.
        # The "Rating" row (no by_size) still renders via fallback.
        if not _valve_row_has_data(fallback, by_size, pipe_sizes):
            continue
        fill = ALT_FILL if i % 2 == 0 else DATA_FILL
        if by_size:
            values = _lvcf_by_size(by_size, pipe_sizes)
            _write_merged_data_row(ws, row, label, values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        else:
            _write_label_value_row(ws, row, label, fallback, col_end=total_cols)
            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).fill = fill
        row += 1

    # === NOTES ===
    # Rendered as a numbered table: col A = position number, col B..end = note text (merged).
    # The flange_type / spectacle_blind strings may reference notes by position ("Note 8,9"),
    # so position numbers must be visible in the sheet.
    if pms.notes:
        _write_section_header(ws, row, "Notes", col_end=total_cols)
        row += 1
        for idx, note in enumerate(pms.notes, start=1):
            # Col 1 = note position number; col 2..end = note text (merged)
            _apply_style(ws, row, 1, font=LABEL_FONT, fill=NOTES_FILL, alignment=CENTER).value = idx
            _apply_style(ws, row, pipe_col_start, font=NOTE_FONT, fill=NOTES_FILL, alignment=LEFT).value = note
            if total_cols > pipe_col_start:
                ws.merge_cells(start_row=row, start_column=pipe_col_start, end_row=row, end_column=total_cols)
            for c in range(pipe_col_start + 1, total_cols + 1):
                ws.cell(row=row, column=c).fill = NOTES_FILL
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.print_area = f"A1:{get_column_letter(total_cols)}{row}"

    # === BRANCH CHART SHEETS ===
    if pms.branch_charts:
        for chart in pms.branch_charts:
            _write_branch_chart_sheet(wb, chart)

    wb.save(output_path)
    logger.info("Excel PMS saved to %s", output_path)
    return output_path


def generate_pms_excel_bytes(pms: PMSResponse) -> bytes:
    """Generate Excel PMS and return as bytes (no disk write)."""
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = pms.piping_class

    # Re-use the same generation logic but save to buffer
    num_pipe_cols = max(len(pms.pipe_data), 1)
    total_cols = max(num_pipe_cols + 1, 19)
    pipe_col_start = 2

    from openpyxl.utils import get_column_letter as gcl
    # Compact per-size data columns — see generate_pms_excel above.
    for i in range(11, total_cols + 1):
        ws.column_dimensions[gcl(i)].width = 6

    # === HEADER BANNER (rows 1-6): compact spec table ===
    row = _write_pms_header(ws, pms, total_cols)

    # P-T Rating
    _write_section_header(ws, row, "Pressure-Temperature Rating", col_end=total_cols)
    row += 1

    pt = pms.pressure_temperature

    # Pressure row (compact label, comma form per the reference sheet)
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=DATA_FILL, alignment=LEFT).value = "Press., barg"
    for i, press in enumerate(pt.pressures):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=DATA_FONT, fill=DATA_FILL).value = press
    # Hydrotest header — single-line, sits to the right of the P/T columns
    ht_col = pipe_col_start + len(pt.temperatures) + 1
    if pms.hydrotest_pressure and ht_col <= total_cols:
        _apply_style(ws, row, ht_col, font=LABEL_FONT, fill=ALT_FILL, alignment=CENTER).value = "Hydrotest Pr. (barg)"
    row += 1

    # Temperature row — single combined row using the labels (e.g. "-29 to 38")
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Temp., °C"
    temp_values = pt.temp_labels if pt.temp_labels else [str(t) for t in pt.temperatures]
    for i, lbl in enumerate(temp_values):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=DATA_FONT, fill=ALT_FILL).value = lbl
    if pms.hydrotest_pressure and ht_col <= total_cols:
        _apply_style(ws, row, ht_col, font=DATA_FONT, fill=DATA_FILL, alignment=CENTER).value = pms.hydrotest_pressure
    row += 1

    # Pipe Data
    _write_section_header(ws, row, "Pipe Data", col_end=total_cols)
    row += 1
    if pms.pipe_code:
        _write_label_value_row(ws, row, "Code", pms.pipe_code, col_end=total_cols)
        row += 1

    pipe_rows = [
        ("Size(in)", [p.size_inch for p in pms.pipe_data]),
        ("O.D.mm", [p.od_mm for p in pms.pipe_data]),
    ]
    if any((getattr(p, "id_mm", 0) or 0) > 0 for p in pms.pipe_data):
        pipe_rows.append(("I.D.mm", [p.id_mm for p in pms.pipe_data]))
    pipe_rows.extend([
        ("Sch.", [_short_schedule(p.schedule) for p in pms.pipe_data]),
        ("WT.mm", [p.wall_thickness_mm for p in pms.pipe_data]),
        ("Type", [p.pipe_type for p in pms.pipe_data]),
        ("MOC", [p.material_spec for p in pms.pipe_data]),
        ("Ends", [p.ends for p in pms.pipe_data]),
    ])
    # Rows that benefit from merging (repeated values across sizes)
    mergeable_pipe_rows = {"Type", "MOC", "Ends"}
    for label, values in pipe_rows:
        is_alt = label in ("Size(in)", "Sch.", "Type", "Ends")
        fill = ALT_FILL if is_alt else DATA_FILL
        if label in mergeable_pipe_rows:
            _write_merged_data_row(ws, row, label, values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        else:
            _apply_style(ws, row, 1, font=LABEL_FONT, fill=fill, alignment=LEFT).value = label
            for i, val in enumerate(values):
                col = pipe_col_start + i
                if col <= total_cols:
                    _apply_style(ws, row, col, font=DATA_FONT, fill=fill).value = val
        row += 1

    # Fittings (SIZE-WISE DATA)
    _write_section_header(ws, row, "Fittings — Butt Weld (SCH to match pipe)", col_end=total_cols)
    row += 1

    # Size columns header — written for column-mapping correctness, then
    # hidden because per spec only Pipe Data shows the size labels visually.
    _apply_style(ws, row, 1, font=LABEL_FONT, fill=ALT_FILL, alignment=LEFT).value = "Size (in)"
    for i, fitting in enumerate(pms.fittings_by_size):
        col = pipe_col_start + i
        if col <= total_cols:
            _apply_style(ws, row, col, font=LABEL_FONT, fill=ALT_FILL, alignment=CENTER).value = fitting.size_inch
    for c in range(pipe_col_start + len(pms.fittings_by_size), total_cols + 1):
        _apply_style(ws, row, c, fill=ALT_FILL)
    ws.row_dimensions[row].hidden = True
    row += 1

    # Type row — full descriptive text from AI (e.g. "Butt Weld (SCH to match pipe), Seamless")
    type_values = [f.fitting_type or "" for f in pms.fittings_by_size]
    _write_merged_data_row(ws, row, "Type", type_values, col_start=pipe_col_start,
                           total_cols=total_cols, fill=DATA_FILL)
    row += 1

    # Rating row (GRE classes only; otherwise skipped).
    if (getattr(pms.fittings, "rating", "") or "").strip():
        _write_label_value_row(ws, row, "Rating", pms.fittings.rating, col_end=total_cols)
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    # Fitting properties — see first renderer for behavior explanation.
    fitting_props = [
        ("MOC", lambda f: f.material_spec),
        ("Elbow", lambda f: f.elbow_standard),
        ("Tee", lambda f: f.tee_standard),
        ("Mold. Tee", lambda f: f.mold_tee_standard),
        ("Red.", lambda f: f.reducer_standard),
        ("Red. Sad", lambda f: f.red_saddle_standard),
        ("Cap", lambda f: f.cap_standard),
        ("Coupl", lambda f: f.coupling_standard),
        ("Plug", lambda f: f.plug_standard),
        ("Union", lambda f: f.union_standard),
        ("Sockolet", lambda f: f.sockolet_standard),
        ("Weldolet", lambda f: f.weldolet_spec),
        ("Adaptor", lambda f: f.adaptor_standard),
        ("Nipple", lambda f: f.nipple_standard),
        ("Swage", lambda f: f.swage_standard),
    ]
    _OPTIONAL_ROWS = {"Coupl", "Union", "Sockolet", "Nipple", "Swage",
                      "Mold. Tee", "Red. Sad", "Adaptor"}

    # Plug is always a small-bore row — threaded plugs only apply at sizes
    # ≤ 1.5" in this project, regardless of class. Even when the AI emits
    # plug_standard on larger sizes (e.g. A40 MOC-split), only the
    # small-bore value is shown; the ≥ 2" portion of the Plug row stays
    # blank.
    fitting_sizes = [f.size_inch for f in pms.fittings_by_size]
    plug_start_col = pipe_col_start
    plug_end_col = _size_column_index(fitting_sizes, "1.5", pipe_col_start=pipe_col_start)

    for prop_idx, (label, getter) in enumerate(fitting_props):
        fill = ALT_FILL if prop_idx % 2 == 0 else DATA_FILL
        prop_values = [getter(f) or "" for f in pms.fittings_by_size]
        if label in _OPTIONAL_ROWS and not _any_nonempty(prop_values):
            continue
        if label == "Plug":
            # Use the first non-empty small-bore value (size ≤ 1.5"); if the
            # AI emitted plug only for large sizes (shouldn't happen but
            # defend anyway), fall back to any non-empty value.
            plug_value = ""
            for f, v in zip(pms.fittings_by_size, prop_values):
                if not v:
                    continue
                try:
                    if float(str(f.size_inch).strip().rstrip('"')) <= 1.5:
                        plug_value = v
                        break
                except (ValueError, TypeError):
                    continue
            if not plug_value:
                plug_value = next((v for v in prop_values if v), "")
            _write_range_value_row(ws, row, label, plug_value,
                                   start_col=plug_start_col, end_col=plug_end_col,
                                   total_cols=total_cols, fill=fill)
        else:
            _write_merged_data_row(ws, row, label, prop_values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        row += 1

    pipe_sizes = [p.size_inch for p in pms.pipe_data]

    # Flange
    _write_section_header(ws, row, "Flange", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1
    has_extra_flange = bool(pms.flange.compact_flange) or bool(pms.flange.hub_connector)
    type_label = "WN Flange" if has_extra_flange else "Type"
    flange_items = [
        ("MOC", pms.flange.material_spec, False),
        ("Face", pms.flange.face_type, False),
        (type_label, pms.flange.flange_type, False),
        ("Standard", pms.flange.standard, False),
    ]
    if has_extra_flange:
        flange_items.append(("Compact Flange", pms.flange.compact_flange, True))
        flange_items.append(("Hub Connector", pms.flange.hub_connector, True))

    offset_col = _size_column_index(pipe_sizes, "3", pipe_col_start=pipe_col_start)
    for i, (lbl, val, offset) in enumerate(flange_items):
        fill = ALT_FILL if i % 2 == 0 else DATA_FILL
        if offset and val:
            _write_label_offset_value_row(ws, row, lbl, val,
                                          value_start_col=offset_col,
                                          col_end=total_cols, fill=fill)
        else:
            _write_label_value_row(ws, row, lbl, val, col_end=total_cols)
            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).fill = fill
        row += 1

    # Spectacle Blind
    _write_section_header(ws, row, "Spectacle Blind / Spacer Blinds", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1
    _write_label_value_row(ws, row, "MOC", pms.spectacle_blind.material_spec, col_end=total_cols)
    row += 1
    # Spectacle row — split at 14" (ASME B 16.48 project cutoff)
    _render_spectacle_blind_row(ws, row, pms, pipe_col_start, total_cols)
    ws.row_dimensions[row].height = 25
    row += 1

    # Bolts/Nuts/Gaskets
    _write_section_header(ws, row, "Bolts / Nuts / Gaskets", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1
    # Washers and Gasket #2 are optional (GRE A50/A51/A52); empty → row hidden.
    bng_rows = [
        ("Stud Bolts", pms.bolts_nuts_gaskets.stud_bolts, False),
        ("Hex Nuts", pms.bolts_nuts_gaskets.hex_nuts, False),
        ("Washers", getattr(pms.bolts_nuts_gaskets, "washers", ""), True),
        ("Gasket", pms.bolts_nuts_gaskets.gasket, False),
        ("Gasket", getattr(pms.bolts_nuts_gaskets, "gasket_2", ""), True),
    ]
    rendered = 0
    for lbl, val, optional in bng_rows:
        if optional and not (val or "").strip():
            continue
        _write_label_value_row(ws, row, lbl, val, col_end=total_cols)
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).fill = ALT_FILL if rendered % 2 == 0 else DATA_FILL
        row += 1
        rendered += 1

    # Valves
    _write_section_header(ws, row, "Valves", col_end=total_cols)
    row += 1
    _write_size_header_row(ws, row, pipe_sizes, col_start=pipe_col_start, total_cols=total_cols)
    row += 1
    valve_types = [
        ("Rating", pms.valves.rating, []),
        ("Ball", pms.valves.ball, pms.valves.ball_by_size),
        ("Gate", pms.valves.gate, pms.valves.gate_by_size),
        ("Globe", pms.valves.globe, pms.valves.globe_by_size),
        ("Check", pms.valves.check, pms.valves.check_by_size),
        ("Butterfly", pms.valves.butterfly, pms.valves.butterfly_by_size),
        ("DBB (Inst)", pms.valves.dbb_inst, pms.valves.dbb_inst_by_size),
        ("DBB", pms.valves.dbb, pms.valves.dbb_by_size),
    ]
    for i, (lbl, fallback, by_size) in enumerate(valve_types):
        # Skip rows with no real data — same logic as the primary renderer.
        if not _valve_row_has_data(fallback, by_size, pipe_sizes):
            continue
        fill = ALT_FILL if i % 2 == 0 else DATA_FILL
        if by_size:
            values = _lvcf_by_size(by_size, pipe_sizes)
            _write_merged_data_row(ws, row, lbl, values, col_start=pipe_col_start,
                                   total_cols=total_cols, fill=fill)
        else:
            _write_label_value_row(ws, row, lbl, fallback, col_end=total_cols)
            for c in range(1, total_cols + 1):
                ws.cell(row=row, column=c).fill = fill
        row += 1

    # Notes — numbered: col 1 = position number, col 2..end = merged text
    if pms.notes:
        _write_section_header(ws, row, "Notes", col_end=total_cols)
        row += 1
        for idx, note in enumerate(pms.notes, start=1):
            _apply_style(ws, row, 1, font=LABEL_FONT, fill=NOTES_FILL, alignment=CENTER).value = idx
            _apply_style(ws, row, pipe_col_start, font=NOTE_FONT, fill=NOTES_FILL, alignment=LEFT).value = note
            if total_cols > pipe_col_start:
                ws.merge_cells(start_row=row, start_column=pipe_col_start, end_row=row, end_column=total_cols)
            for c in range(pipe_col_start + 1, total_cols + 1):
                ws.cell(row=row, column=c).fill = NOTES_FILL
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # === BRANCH CHART SHEETS ===
    if pms.branch_charts:
        for chart in pms.branch_charts:
            _write_branch_chart_sheet(wb, chart)

    wb.save(buf)
    return buf.getvalue()


def _write_branch_chart_sheet(wb, chart):
    """Write a branch connection chart as a separate Excel sheet."""
    ws = wb.create_sheet(title=f"Chart {chart.chart_id}")

    # Color map for connection types
    FILL_T = PatternFill("solid", fgColor="C6EFCE")    # Green — Tee
    FILL_W = PatternFill("solid", fgColor="FFC7CE")    # Red — Weldolet
    FILL_H = PatternFill("solid", fgColor="FFEB9C")    # Yellow — Threadolet
    FILL_S = PatternFill("solid", fgColor="B4C6E7")    # Blue — Sockolet
    FILL_RT = PatternFill("solid", fgColor="D9E2F3")   # Light blue — Reducing Tee
    FILL_DASH = PatternFill("solid", fgColor="D9D9D9")  # Grey — N/A
    FILL_EMPTY = PatternFill("solid", fgColor="F2F2F2")  # Light grey — empty
    FILL_MAP = {"T": FILL_T, "W": FILL_W, "H": FILL_H, "S": FILL_S, "RT": FILL_RT, "-": FILL_DASH}

    CHART_FONT = Font(name="Arial", bold=True, size=9)
    CELL_FONT = Font(name="Arial", bold=True, size=9)
    HEADER_CELL = Font(name="Arial", bold=True, color="FFFFFF", size=9)

    run_sizes = chart.run_sizes
    branch_sizes = chart.branch_sizes

    # Title row
    row = 1
    ws.cell(row=row, column=1).value = f"BRANCH CONNECTION CHART {chart.chart_id} — {chart.title}"
    ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, color="1F4E79", size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(branch_sizes) + 2)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.cell(row=row, column=1).value = "Branch Table as per API RP 14E"
    ws.cell(row=row, column=1).font = Font(name="Arial", italic=True, size=9, color="666666")
    row += 1

    # Column header: "RUN \\ BRANCH" then branch sizes
    ws.cell(row=row, column=1).value = "RUN \\ BRANCH"
    ws.cell(row=row, column=1).font = HEADER_CELL
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.column_dimensions["A"].width = 14

    for j, bs in enumerate(branch_sizes):
        col = j + 2
        cell = ws.cell(row=row, column=col)
        cell.value = bs
        cell.font = HEADER_CELL
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 7
    row += 1

    # Data rows
    for i, rs in enumerate(run_sizes):
        # Run size header
        cell = ws.cell(row=row, column=1)
        cell.value = rs
        cell.font = CHART_FONT
        cell.fill = SECTION_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

        # Grid cells
        grid_row = chart.grid[i] if i < len(chart.grid) else []
        for j in range(len(branch_sizes)):
            col = j + 2
            val = grid_row[j] if j < len(grid_row) else ""
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.font = CELL_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.fill = FILL_MAP.get(val, FILL_EMPTY)
        row += 1

    row += 1

    # Legend
    ws.cell(row=row, column=1).value = "LEGEND"
    ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    row += 1
    for code, desc in chart.legend.items():
        ws.cell(row=row, column=1).value = code
        ws.cell(row=row, column=1).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).fill = FILL_MAP.get(code, DATA_FILL)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=2).font = Font(name="Arial", size=9)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
